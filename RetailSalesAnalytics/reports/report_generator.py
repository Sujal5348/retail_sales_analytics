import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from config import Config
from analytics.eda_engine import get_monthly_sales_trend, get_category_performance, get_regional_sales, get_top_products

def generate_excel_report():
    """
    Generates an automated executive retail report in Excel format with styled headers and multiple tabs.
    """
    os.makedirs(Config.EXPORT_FOLDER, exist_ok=True)
    filepath = os.path.join(Config.EXPORT_FOLDER, 'Retail_Sales_Executive_Report.xlsx')
    
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="0F172A")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    reports = [
        ('Monthly Sales Trends', get_monthly_sales_trend()),
        ('Category Performance', get_category_performance()),
        ('Regional Breakdown', get_regional_sales()),
        ('Top 10 Products', get_top_products(10))
    ]

    for sheet_title, df in reports:
        ws = wb.create_sheet(title=sheet_title)
        
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Retail Analytics Report - {sheet_title}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(vertical='center')
        ws.row_dimensions[1].height = 30

        ws.append([])

        headers = list(df.columns)
        ws.append(headers)
        header_row_idx = 3
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row_idx, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[header_row_idx].height = 24

        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
            current_row = ws.max_row
            ws.row_dimensions[current_row].height = 20
            for col_idx in range(1, len(row) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                val = cell.value
                if isinstance(val, (int, float)):
                    if 'Revenue' in headers[col_idx-1] or 'Profit' in headers[col_idx-1]:
                        cell.number_format = '$#,##0.00'
                    elif 'Percent' in headers[col_idx-1] or 'Growth' in headers[col_idx-1]:
                        cell.number_format = '0.00%'

        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(filepath)
    print(f"Report exported to: {filepath}")
    return filepath

if __name__ == '__main__':
    generate_excel_report()
